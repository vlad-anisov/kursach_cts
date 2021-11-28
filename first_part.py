from openpyxl import load_workbook
from xltpl.writerx import BookWriter
import erlang
import formulas


def first_part(cities):
    data = get_data(cities)
    write_cities_data(data)
    write_erlang_data()


def get_data(cities):
    data = []
    for city in cities:
        data.append(get_city_data(city))
    return data


def write_erlang_data():
    xl_model = formulas.ExcelModel().loads("test_1.xlsx").finish()
    xl_model.calculate()
    xl_model.write(dirpath="calculate")
    wb = load_workbook("test_1.xlsx")
    sheet = wb.worksheets[0]
    wb2 = load_workbook("calculate/test_1.xlsx")
    sheet2 = wb2.worksheets[0]
    for i in range(1, sheet.max_row + 1):
        if i in [11, 15]:
            for j in range(3, 26, 2):
                erlang_value = float(sheet2.cell(i - 1, j).value)
                sheet.cell(row=i, column=j).value = erlang.extended_b_lines(erlang_value, 0.01)
        if i in [3, 6]:
            for j in range(3, 27):
                sheet.cell(row=i, column=j).value = str(sheet2.cell(i, j).value).replace('.', ',')
    wb.save("first_table.xlsx")


def get_city_data(city):
    data = [city]
    wb = load_workbook("cities.xlsx")
    sheet = wb.worksheets[0]
    for i in range(1, sheet.max_row + 1):
        value = sheet.cell(i, 1).value
        if value and value.lower() == f'г.{city}'.lower():
            data.append(0)
            for j in range(1, 10):
                if sheet.cell(i - j, 1).value == 'городское население':
                    value = str(sheet.cell(i - j, 4).value / 1000).replace('.', ',')
                    data[1] = value
                    break
            data.append(0)
            for j in range(1, 10):
                if sheet.cell(i + j, 1).value == 'сельское население':
                    value = str(sheet.cell(i + j, 4).value / 1000).replace('.', ',')
                    data[2] = value
                    break
            break
    return data


def write_cities_data(data_of_cities):
    rows = list(map(list, zip(*data_of_cities)))
    cities = rows[0]
    population_sizes = [float(item.replace(",", ".")) for sublist in list(zip(rows[1], rows[2])) for item in sublist]
    context = {
        "cities": cities,
        "population_sizes": population_sizes,
        "norms_of_telephone_density": get_norms_of_telephone_density(population_sizes),
        "number_of_subscribers": get_number_of_subscribers(population_sizes),
    }
    create_xlsx_report(context, "template_test_1.xlsx")


def create_xlsx_report(context, template_path):
    """Creating xlsx report"""
    writer = BookWriter(template_path)
    context["sheet_name"] = "sheet"
    writer.render_book(payloads=[context])
    writer.save('test_1.xlsx')


def get_norms_of_telephone_density(population_sizes):
    norms_of_telephone_density = []
    for index, size in enumerate(population_sizes, 1):
        if index % 2 == 0:
            norms_of_telephone_density.append(150)
        else:
            if 1 <= size < 10:
                norms_of_telephone_density.append(215)
            elif 10 <= size < 20:
                norms_of_telephone_density.append(255)
            elif 20 <= size < 50:
                norms_of_telephone_density.append(290)
            elif 50 <= size < 100:
                norms_of_telephone_density.append(320)
            elif 100 <= size < 500:
                norms_of_telephone_density.append(370)
            else:
                norms_of_telephone_density.append(415)
    return norms_of_telephone_density


def get_number_of_subscribers(population_sizes):
    norms_of_telephone_density = []
    for index, size in enumerate(population_sizes, 1):
        if index % 2 == 0:
            norms_of_telephone_density.append(0.12)
        else:
            if 1 <= size < 10:
                norms_of_telephone_density.append(0.7)
            elif 10 <= size < 20:
                norms_of_telephone_density.append(0.7)
            elif 20 <= size < 50:
                norms_of_telephone_density.append(0.6)
            elif 50 <= size < 100:
                norms_of_telephone_density.append(0.5)
            elif 100 <= size < 500:
                norms_of_telephone_density.append(0.4)
            else:
                norms_of_telephone_density.append(0.3)
    return norms_of_telephone_density
